import os
import mysql.connector
from dotenv import load_dotenv
from openpyxl import load_workbook

# Load DB credentials from .env
load_dotenv()
DB_HOST = os.getenv("DB_HOST")
DB_USER = os.getenv("DB_USER")
DB_PASSWORD = os.getenv("DB_PASSWORD")

def load_excel_data(file_path: str, sheet_name: str) -> list:
    workbook = load_workbook(filename=file_path)
    worksheet = workbook[sheet_name]

    data = []
    for i in range(2, worksheet.max_row + 1):
        row_data = []
        for j in range(1, worksheet.max_column + 1):
            row_data.append(worksheet.cell(row=i, column=j).value)
        data.append(row_data)
    return data

def insert_data_to_mysql(
    data: list,
    database_name: str,
    table_name: str,
    db_host: str = DB_HOST,
    db_user: str = DB_USER,
    db_password: str = DB_PASSWORD
) -> bool:
    connection = mysql.connector.connect(
        host=db_host,
        user=db_user,
        password=db_password,
        database=database_name
    )
    cursor = connection.cursor()

    values_str = ", ".join(str(tuple(row)).replace("None", "NULL") for row in data)
    insert_query = f"INSERT INTO {table_name} VALUES {values_str}"

    try:
        cursor.execute(insert_query)
        connection.commit()
        return True
    except mysql.connector.Error:
        connection.rollback()
        return False
    finally:
        cursor.close()
        connection.close()

def run_import(
    file_path: str,
    sheet_name: str,
    database_name: str,
    table_name: str
) -> bool:
    data = load_excel_data(file_path, sheet_name)
    return insert_data_to_mysql(data, database_name, table_name)

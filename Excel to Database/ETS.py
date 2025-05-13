# importing Modules
from openpyxl import load_workbook
import os
from time import sleep
from pyautogui import press, typewrite

# input values
FilePath = f"{os.getcwd()}\\Downloads\\Book1.xlsx"
SheetName = "Sheet1"


# caution: The excel sheet must have more than 1 row and 1 column.

# cursor creation
workbook = load_workbook (filename = FilePath)
wb = workbook[SheetName]

# Defining Funtions and Main Program
def GetData():
    data = []
    for i in range (1, wb.max_row + 1):
        data2 = []
        for j in range (1, wb.max_column+1):
            data2.append(wb.cell(row=i, column=j).value)

        data.append(data2)

    return data

def Enter1(data):
    for i in data:
        if i[1] != 0:
            print (i)
            typewrite(i[0].split("_")[0])
            sleep(5)
            press("tab")
            typewrite(str(i[1]))
            sleep(1.8)
            press("enter")
            sleep(3.5)


def Run():
    print ("Program Started")
    sleep(3)
    data = GetData()
    print (data, "\n")
    Enter1(data)

Run()